from scrapy.spiders import Spider
from scrapy.http import Request, FormRequest

from parse_linkedin.items import Person

from openpyxl import load_workbook


class LinkedinBioSpider(Spider):
    def __init__(self, *a, **kw):
        super(LinkedinBioSpider, self).__init__(*a, **kw)
        wb = load_workbook('names.xlsx')
        ws = wb.active
        self.names = []
        for row in ws.iter_rows():
            self.names.append(row[0].value)

    name = "parse_linkedin"
    allowed_domains = ["www.linkedin.com"]
    start_urls = [
        "http://www.linkedin.com/"
    ]

    def parse(self, response):
        for name in self.names:
            first, last = name.split()
            yield FormRequest.from_response(response, formxpath='//div[@class="footer"]//form',\
            formdata={'first':first, 'last':last}, callback=self.parsePerson, meta = {'name':name})

    def parsePerson(self,response):
        find_person = response.css('div.profile-card h3 a::attr(href)')
        if len(find_person) > 0:
            yield Request(find_person[0].extract(), callback = self.parsePerson, meta = response.meta)
        else:
            person = Person()
            person['name'] = response.meta['name']
            person['bio'] = response.css('section#summary p').extract()[0]
            yield person