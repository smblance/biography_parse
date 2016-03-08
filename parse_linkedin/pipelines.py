# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html

from openpyxl import load_workbook

class SaveToXLSX(object):

	def __init__(self):
		self.wb = load_workbook('names.xlsx')

	def process_item(self, item, spider):
		ws = self.wb.active
		for row in ws.iter_rows():
			if row[0].value == item['name']:
				ws.cell(coordinate = 'B' + str(row[0].row)).value = item['bio']
				break
		# self.wb.active.append([item['name'],item['bio']])
		self.wb.save('names.xlsx')

		return item
